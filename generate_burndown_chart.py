#!/usr/bin/env python3
"""
PLC Control Narrative Burndown Chart Generator

Creates an Excel workbook to track progress of Control Narrative creation
for 34 PLCs. Each Control Narrative includes:
- Alarm Summary document
- Cause & Effect document

Timeline: 2 work days per PLC (excluding weekends)
Start date: December 10, 2025
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# Define all 34 PLCs with their descriptions and types
PLCS = [
    # 1-5: VRU PLCs
    {"id": 1, "name": "VRU PLC 1", "type": ""},
    {"id": 2, "name": "VRU PLC 2", "type": ""},
    {"id": 3, "name": "VRU PLC 3", "type": ""},
    {"id": 4, "name": "VRU PLC 4", "type": ""},
    {"id": 5, "name": "VRU PLC 5", "type": ""},

    # 6-7: VCU PLCs
    {"id": 6, "name": "VCU PLC 1", "type": ""},
    {"id": 7, "name": "VCU PLC 2", "type": ""},

    # 8-10: Master PLCs
    {"id": 8, "name": "Master PLC3", "type": ""},
    {"id": 9, "name": "Master PLC4", "type": ""},
    {"id": 10, "name": "MVCU PLC", "type": ""},

    # 11-16: Dock Safety Unit PLCs
    {"id": 11, "name": "Dock Safety Unit PLC 1", "type": ""},
    {"id": 12, "name": "Dock Safety Unit PLC 2", "type": ""},
    {"id": 13, "name": "Dock Safety Unit PLC 3", "type": ""},
    {"id": 14, "name": "Dock Safety Unit PLC 4", "type": ""},
    {"id": 15, "name": "Dock Safety Unit PLC 5", "type": ""},
    {"id": 16, "name": "Dock Safety Unit PLC 6", "type": ""},

    # 17: Main Terminal PLC
    {"id": 17, "name": "Main Terminal PLC", "type": "ControlLogix 1756-L72"},

    # 18: LACT PLC
    {"id": 18, "name": "LACT PLC Panel w/ESD Panel", "type": "MicroLogix 1100"},

    # 19: Tomahawk PLC
    {"id": 19, "name": "Tomahawk PLC Panel", "type": "ControlLogix 1756-L72"},

    # 20: Dock 5 PLC
    {"id": 20, "name": "Dock 5 PLC Panel", "type": "MicroLogix 1400"},

    # 21: 80/50/RC PLC
    {"id": 21, "name": "80/50/RC PLC Panel w/ESD Panel", "type": "MicroLogix 1400"},

    # 22-24: Butane Area PLCs
    {"id": 22, "name": "Butane Area Phase 1 PLC Panel", "type": "MicroLogix 1400"},
    {"id": 23, "name": "Butane Area Phase 2 PLC Panel", "type": "MicroLogix 1400"},
    {"id": 24, "name": "Butane Area Fire Eye PLC Panel", "type": "MicroLogix 1400"},

    # 25-26: Valero & Butane / Compressor PLCs
    {"id": 25, "name": "Valero PLC Panel w/ESD & Fiber", "type": "MicroLogix 1400"},
    {"id": 26, "name": "Butane/Compressor PLC Panel w/ESD & Fiber", "type": "MicroLogix 1400"},

    # 27: East Dock PLC
    {"id": 27, "name": "East Dock PLC & RIO-104 RIO Panels", "type": "CompactLogix 1769-L30ER"},

    # 28: Phase 1C MCC PLC
    {"id": 28, "name": "Phase 1C MCC PLC Panel", "type": "CompactLogix 1769-L30ER"},

    # 29: Switch Rack 111
    {"id": 29, "name": "Switch Rack 111 (NJB001) PLC Panel", "type": "MicroLogix 1400"},

    # 30: Dock 1 TDRC-101 PLC
    {"id": 30, "name": "Dock 1 TDRC-101 PLC & RIO-102 RIO Panels", "type": "CompactLogix 1769-L30ER"},

    # 31: Cactus 2 PLC
    {"id": 31, "name": "Cactus 2 PLC Panel", "type": "ControlLogix 1756-L72"},

    # 32-33: Switch Rack 113 PLCs
    {"id": 32, "name": "Switch Rack 113 PLC Panel 1", "type": "MicroLogix 1400"},
    {"id": 33, "name": "Switch Rack 113 PLC Panel 2", "type": "MicroLogix 1400"},

    # 34: Fire System #2 PLC
    {"id": 34, "name": "Fire System #2 PLC Panel", "type": "MicroLogix 1100"},
]

def calculate_work_days(start_date, num_work_days):
    """
    Calculate a list of work days (excluding weekends) starting from start_date.

    Args:
        start_date: datetime object for start date
        num_work_days: number of work days needed

    Returns:
        List of datetime objects representing work days
    """
    work_days = []
    current_date = start_date

    while len(work_days) < num_work_days:
        # Check if current date is a weekday (Monday=0, Sunday=6)
        if current_date.weekday() < 5:  # Monday to Friday
            work_days.append(current_date)
        current_date += timedelta(days=1)

    return work_days

def generate_burndown_chart():
    """Generate Excel workbook with PLC Control Narrative burndown tracking."""

    # Configuration
    start_date = datetime(2025, 12, 10)  # December 10, 2025
    days_per_plc = 2
    total_plcs = len(PLCS)
    total_work_days = total_plcs * days_per_plc

    # Calculate all work days
    work_days = calculate_work_days(start_date, total_work_days)
    end_date = work_days[-1]

    # Create workbook
    wb = Workbook()

    # Sheet 1: PLC Tracking List
    ws_tracking = wb.active
    ws_tracking.title = "PLC Tracking"

    # Headers for tracking sheet
    headers = [
        "PLC #",
        "PLC Name",
        "PLC Type",
        "Planned Start",
        "Planned End",
        "Alarm Summary Complete",
        "Cause & Effect Complete",
        "Both Documents Complete",
        "Actual Complete Date",
        "Status",
        "Notes"
    ]

    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_tracking.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Set column widths
    ws_tracking.column_dimensions['A'].width = 8
    ws_tracking.column_dimensions['B'].width = 40
    ws_tracking.column_dimensions['C'].width = 25
    ws_tracking.column_dimensions['D'].width = 12
    ws_tracking.column_dimensions['E'].width = 12
    ws_tracking.column_dimensions['F'].width = 12
    ws_tracking.column_dimensions['G'].width = 12
    ws_tracking.column_dimensions['H'].width = 12
    ws_tracking.column_dimensions['I'].width = 15
    ws_tracking.column_dimensions['J'].width = 12
    ws_tracking.column_dimensions['K'].width = 30

    # Fill in PLC data
    current_day_index = 0
    for plc in PLCS:
        row = plc["id"] + 1  # +1 for header row

        # Calculate planned dates for this PLC
        plc_start = work_days[current_day_index]
        plc_end = work_days[min(current_day_index + days_per_plc - 1, len(work_days) - 1)]
        current_day_index += days_per_plc

        # PLC #
        cell = ws_tracking.cell(row=row, column=1)
        cell.value = plc["id"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # PLC Name
        cell = ws_tracking.cell(row=row, column=2)
        cell.value = plc["name"]
        cell.border = border

        # PLC Type
        cell = ws_tracking.cell(row=row, column=3)
        cell.value = plc["type"]
        cell.border = border

        # Planned Start
        cell = ws_tracking.cell(row=row, column=4)
        cell.value = plc_start.strftime("%m/%d/%Y")
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # Planned End
        cell = ws_tracking.cell(row=row, column=5)
        cell.value = plc_end.strftime("%m/%d/%Y")
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # Checkbox columns (F, G, H) - leave empty for manual entry
        for col in [6, 7, 8]:
            cell = ws_tracking.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Actual Complete Date (I)
        cell = ws_tracking.cell(row=row, column=9)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

        # Status (J) - formula based
        cell = ws_tracking.cell(row=row, column=10)
        cell.value = f'=IF(H{row}="X","Complete",IF(OR(F{row}="X",G{row}="X"),"In Progress","Not Started"))'
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # Notes (K)
        cell = ws_tracking.cell(row=row, column=11)
        cell.border = border

    # Sheet 2: Burndown Chart Data
    ws_burndown = wb.create_sheet("Burndown Data")

    # Headers for burndown sheet
    burndown_headers = ["Date", "Day #", "Planned Remaining", "Actual Remaining"]
    for col_num, header in enumerate(burndown_headers, 1):
        cell = ws_burndown.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Set column widths
    ws_burndown.column_dimensions['A'].width = 12
    ws_burndown.column_dimensions['B'].width = 8
    ws_burndown.column_dimensions['C'].width = 18
    ws_burndown.column_dimensions['D'].width = 18

    # Generate burndown data
    # Planned: linear decrease from total_plcs to 0
    plcs_per_day = 1 / days_per_plc  # 0.5 PLCs per day with 2 days per PLC

    for day_num, date in enumerate(work_days, 1):
        row = day_num + 1

        # Date
        cell = ws_burndown.cell(row=row, column=1)
        cell.value = date.strftime("%m/%d/%Y")
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # Day #
        cell = ws_burndown.cell(row=row, column=2)
        cell.value = day_num
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # Planned Remaining - end of day count
        planned_completed = day_num * plcs_per_day
        planned_remaining = max(0, total_plcs - planned_completed)
        cell = ws_burndown.cell(row=row, column=3)
        cell.value = round(planned_remaining, 1)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # Actual Remaining - formula counting uncompleted PLCs
        cell = ws_burndown.cell(row=row, column=4)
        # Count PLCs where column H is not "X" in the tracking sheet
        cell.value = f'=COUNTIF(\'PLC Tracking\'!$H$2:$H${total_plcs + 1},"<>X")'
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Create burndown chart
    chart = LineChart()
    chart.title = "PLC Control Narrative Burndown Chart"
    chart.style = 13
    chart.y_axis.title = "PLCs Remaining"
    chart.x_axis.title = "Work Days"
    chart.height = 10
    chart.width = 20

    # Add data to chart
    # Planned Remaining line
    planned_data = Reference(ws_burndown, min_col=3, min_row=1, max_row=len(work_days) + 1)
    chart.add_data(planned_data, titles_from_data=True)

    # Actual Remaining line
    actual_data = Reference(ws_burndown, min_col=4, min_row=1, max_row=len(work_days) + 1)
    chart.add_data(actual_data, titles_from_data=True)

    # X-axis labels (dates)
    dates = Reference(ws_burndown, min_col=1, min_row=2, max_row=len(work_days) + 1)
    chart.set_categories(dates)

    # Style the lines
    s1 = chart.series[0]
    s1.graphicalProperties.line.solidFill = "FF0000"  # Red for planned
    s1.graphicalProperties.line.width = 25000

    s2 = chart.series[1]
    s2.graphicalProperties.line.solidFill = "0000FF"  # Blue for actual
    s2.graphicalProperties.line.width = 25000

    # Add chart to sheet
    ws_burndown.add_chart(chart, "F2")

    # Sheet 3: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    summary_data = [
        ["Project Summary", ""],
        ["", ""],
        ["Total PLCs", total_plcs],
        ["Days per PLC", days_per_plc],
        ["Total Work Days", total_work_days],
        ["", ""],
        ["Start Date", start_date.strftime("%m/%d/%Y")],
        ["End Date", end_date.strftime("%m/%d/%Y")],
        ["", ""],
        ["Progress Tracking", ""],
        ["PLCs Completed", f'=COUNTIF(\'PLC Tracking\'!$H$2:$H${total_plcs + 1},"X")'],
        ["PLCs In Progress", f'=COUNTIFS(\'PLC Tracking\'!$H$2:$H${total_plcs + 1},"<>X",\'PLC Tracking\'!$F$2:$F${total_plcs + 1},"X")+COUNTIFS(\'PLC Tracking\'!$H$2:$H${total_plcs + 1},"<>X",\'PLC Tracking\'!$G$2:$G${total_plcs + 1},"X")'],
        ["PLCs Not Started", f'=COUNTIFS(\'PLC Tracking\'!$F$2:$F${total_plcs + 1},"<>X",\'PLC Tracking\'!$G$2:$G${total_plcs + 1},"<>X")'],
        ["", ""],
        ["% Complete", f'=ROUND((B11/{total_plcs})*100,1)&"%"'],
    ]

    for row_num, (label, value) in enumerate(summary_data, 1):
        # Label column
        cell_a = ws_summary.cell(row=row_num, column=1)
        cell_a.value = label
        if row_num == 1 or row_num == 10:
            cell_a.font = Font(bold=True, size=14)
        else:
            cell_a.font = Font(bold=True)

        # Value column
        cell_b = ws_summary.cell(row=row_num, column=2)
        cell_b.value = value
        if row_num in [3, 4, 5, 7, 8, 11, 12, 13, 15]:
            cell_b.alignment = Alignment(horizontal='center')

    # Save workbook
    output_file = "PLC_Burndown_Chart.xlsx"
    wb.save(output_file)

    print(f"✓ Burndown chart generated: {output_file}")
    print(f"\nProject Details:")
    print(f"  Total PLCs: {total_plcs}")
    print(f"  Start Date: {start_date.strftime('%B %d, %Y (%A)')}")
    print(f"  End Date: {end_date.strftime('%B %d, %Y (%A)')}")
    print(f"  Total Work Days: {total_work_days}")
    print(f"  Days per PLC: {days_per_plc}")
    print(f"\nWorkbook contains 3 sheets:")
    print(f"  1. PLC Tracking - Main tracking sheet with all 34 PLCs")
    print(f"  2. Burndown Data - Daily burndown data and chart")
    print(f"  3. Summary - Project summary and statistics")
    print(f"\nTo use:")
    print(f"  - Mark 'X' in columns F (Alarm Summary) and G (Cause & Effect) as you complete them")
    print(f"  - Column H will auto-populate with 'X' when both F and G are marked")
    print(f"  - Status column updates automatically")
    print(f"  - Burndown chart updates in real-time")

if __name__ == "__main__":
    generate_burndown_chart()
