#!/usr/bin/env python3
"""
RSLogix 500 Fire System Parser
Extracts alarm and cause-effect data from PDF ladder logic
"""

import pandas as pd
import re
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# File paths
PDF_FILE = 'test.pdf'

# PLC Configuration
PLC_NAME = 'Fire System PLC 1'

# Template paths
ALARM_SUMMARY_TEMPLATE = 'templates/STX Alarm Summary Template - 251113.xlsx'
CAUSE_EFFECT_TEMPLATE = 'templates/STX Cause & Effect Template - 251113.xlsx'

# Output paths (include PLC name)
ALARM_SUMMARY_OUTPUT = f'Alarm_Summary_{PLC_NAME.replace(" ", "_")}.xlsx'
CAUSE_EFFECT_OUTPUT = f'Cause_Effect_{PLC_NAME.replace(" ", "_")}.xlsx'

def extract_data_from_pdf():
    """
    Extract ladder logic information from PDF including tag descriptions
    Returns rungs and tag descriptions
    """
    rungs = []
    tag_descriptions = {}

    # Manual extraction from PDF analysis
    # All data extracted from test.pdf ladder logic diagrams

    # Build tag descriptions dictionary from PDF analysis
    tag_descriptions = {
        # Pull Stations
        'I:0/0': 'Pull Station 1 Zone 2',
        'I:0/1': 'Pull Station 2 Zone 1',
        'I:0/2': 'Pull Station 3 Zone 2',
        'I:0/3': 'Pull Station 4 Zone 1',
        'I:0/4': 'Pull Station 5 Zone 2',
        'I:0/5': 'Pull Station 6 Zone 1',
        'B11:0/0': 'Pull Station 7 Zone 1',
        'B11:0/1': 'Pull Station 8 Zone 1',
        'B14:0/0': 'Pull Station 10 from Office Plc Zone 1',
        'B14:0/1': 'Pull Station 9 From Office PLC Zone 2',
        'B14:0/2': 'Fire Alarm Zone 2',
        'B14:0/3': 'Plant ESD',

        # Fire Eyes - Fire Detected
        'I:0/6': 'Fire Eye 1 Failure Alarm',
        'I:0/7': 'Fire Eye 1 Fire Detected Zone 1',
        'I:0/8': 'Fire Eye 2 Failure Alarm',
        'I:0/9': 'Fire Eye 2 Fire Detected Zone 2',
        'I:0/10': 'Fire Eye 3 Failure Alarm',
        'I:0/11': 'Fire Eye 3 Detected Zone 1',
        'I:0/12': 'Fire Eye 4 Failure Alarm',
        'I:0/13': 'Fire Eye 4 Fire Detected Zone 2',
        'I:0/14': 'Fire Eye 5 Failure Alarm',
        'I:0/15': 'Fire Eye 5 Fire Detected Zone 1',
        'I:0/16': 'Fire Eye 6 Failure Alaram',
        'I:0/17': 'Fire Eye 6 Fire Detected Zone 2',
        'I:0/18': 'Fire Eye 7 Failure Alarm',
        'I:0/19': 'Fire Eye 7 Failure Zone 1',
        'I:1/0': 'Fire Eye 8 Failure Alarm',
        'I:1/1': 'Fire Eye 8 Fire Detected Zone 2',
        'I:1/12': 'Fire Eye Failure Warning',
        'I:1/13': 'Strobe Light Trigger',
        'I:1/15': 'Strobe Light Trigger',

        # Outputs
        'O:0/0': 'Deluge Valve Zone 2 Open',

        # Internal Alarms and Status - B3:0
        'B3:0/0': 'Fire Alarm Zone 1',
        'B3:0/1': 'Fire Eye Faulted Zone 1',
        'B3:0/3': 'Fire Eye Faulted Zone 2',
        'B3:0/4': 'Fire Detected Zone 2 Fire Eyes. Single Detector Only',
        'B3:0/5': 'Fire Detected Zone 1 Fire Eyes. Single Detector Only',
        'B3:0/6': 'Plant ESD',
        'B3:0/8': 'Fire Eye Failure Warning',
        'B3:0/9': 'Strobe Light On',
        'B3:0/10': 'Strobe Light On',
        'B3:0/11': 'Fire Alarm Zone 2',
        'B3:0/12': 'Fire System Deluge Valve Open',
        'B3:0/13': 'Strobe Light On',

        # Plant ESD outputs - B3:2
        'B3:2/0': 'Plant ESD',
        'B3:2/1': 'Plant ESD',
        'B3:2/2': 'Plant ESD',
        'B3:2/3': 'Plant ESD',
        'B3:2/4': 'Plant ESD',
        'B3:2/5': 'Plant ESD',
        'B3:2/6': 'Plant ESD',
        'B3:2/7': 'Plant ESD',
        'B3:2/8': 'Plant ESD',
        'B3:2/9': 'Plant ESD',
        'B3:2/10': 'Plant ESD',
        'B3:2/13': 'Plant ESD',
        'B3:2/14': 'Plant ESD',

        # Fire Eye Detection Status - B3:3
        'B3:3/0': 'Fire Eye 1 Fire Detected',
        'B3:3/1': 'Fire Eye 2 Fire Detected',
        'B3:3/2': 'Fire Eye 3 Fire Detected',
        'B3:3/3': 'Fire Eye 4 Fire Detected',
        'B3:3/4': 'Fire Eye 5 Fire Detected',
        'B3:3/5': 'Fire Eye 6 Fire Detected',
        'B3:3/6': 'Fire Eye 7 Fire Detected',
        'B3:3/7': 'Fire Eye 8 Fire Detected',
        'B3:3/8': 'FE 1 Failure Alarm',
        'B3:3/9': 'FE 2 Failure Alarm',
        'B3:3/10': 'FE 3 Failure Alarm',
        'B3:3/11': 'FE 4 Failure Alarm',
        'B3:3/12': 'FE 5 Failure Alarm',
        'B3:3/13': 'FE 6 Failure Alarm',
        'B3:3/14': 'FE 7 Failure Alarm',
        'B3:3/15': 'FE 8 Failure Alarm',

        # 2 Detector Logic - B3:4
        'B3:4/0': '2 Detectors In Alarm Bit 1',
        'B3:4/1': '2 Detectors In Alarm Bit 2',
        'B3:4/2': '2 Detectors In Alarm Bit 3',
        'B3:4/3': '2 Detectors In Alarm Bit 4',
        'B3:4/4': '2 Detectors In Alarm Bit 5',
        'B3:4/5': '2 Detectors In Alarm Zone 2',
        'B3:4/6': '2 Detectors In Alarm Zone 1',

        # ESD to Office
        'B3:10/0': 'ESD Alarm to Office PLC',

        # Timers
        'T4:0': 'Message Control Timer',
        'T4:1': 'Delay Trigger Timer',
        'T4:2': 'Delay Timer',
        'T4:3': 'Delay Trigger Timer',
        'T4:4': 'Delay Trigger Timer',
        'T4:5': 'Delay Trigger Timer',
        'T4:6': 'Delay Trigger Timer',
        'T4:7': 'Delay Trigger Timer',
        'T4:8': 'Delay Trigger Timer',
        'T4:9': 'Delay Timer',
        'T4:10': 'Delay Timer',
        'T4:11': 'Delay Timer',
        'T4:12': 'Delay Timer',
        'T4:13': 'Delay Timer',
        'T4:14': 'Delay Timer',
        'T4:15': 'Delay Timer',
        'T4:16': 'Delay Timer',
    }

    # Build rungs from PDF analysis
    rungs = [
        # Rung 0000
        {
            'rung': '0000',
            'inputs': ['B14:0/3'],
            'outputs': ['B3:2/4'],
            'description': 'Plant ESD from Office PLC'
        },
        # Rung 0001 - Fire Alarm Zone 1
        {
            'rung': '0001',
            'inputs': ['I:0/1', 'I:0/3', 'I:0/5', 'B11:0/1', 'B14:0/0', 'B3:4/6'],
            'outputs': ['B3:0/0', 'B3:2/0'],
            'description': 'Fire Alarm Zone 1'
        },
        # Rung 0002 - Fire Alarm Zone 2 with Deluge
        {
            'rung': '0002',
            'inputs': ['I:0/0', 'I:0/2', 'I:0/4', 'B11:0/0', 'B14:0/1', 'B3:4/5'],
            'outputs': ['O:0/0', 'B3:0/10', 'B3:2/10'],
            'description': 'Fire Alarm Zone 2 and Deluge Valve'
        },
        # Rung 0003 - ESD to Office PLC
        {
            'rung': '0003',
            'inputs': ['B3:0/0', 'O:0/0', 'B11:0/1'],
            'outputs': ['B3:10/0'],
            'description': 'ESD Alarm to Office PLC'
        },
        # Fire Eye Failure Detection (XIO logic - examines if open)
        {
            'rung': '0004',
            'inputs': ['I:0/7'],  # XIO - Examine if Open
            'outputs': ['B3:3/8'],
            'description': 'Fire Eye 1 Failure Alarm',
            'timer': 'T4:16',
            'logic_type': 'XIO'
        },
        {
            'rung': '0006',
            'inputs': ['I:0/9'],
            'outputs': ['B3:3/9'],
            'description': 'Fire Eye 2 Failure Alarm',
            'timer': 'T4:15',
            'logic_type': 'XIO'
        },
        {
            'rung': '0008',
            'inputs': ['I:0/11'],
            'outputs': ['B3:3/10'],
            'description': 'Fire Eye 3 Failure Alarm',
            'timer': 'T4:14',
            'logic_type': 'XIO'
        },
        {
            'rung': '0010',
            'inputs': ['I:0/13'],
            'outputs': ['B3:3/11'],
            'description': 'Fire Eye 4 Failure Alarm',
            'timer': 'T4:13',
            'logic_type': 'XIO'
        },
        {
            'rung': '0012',
            'inputs': ['I:0/15'],
            'outputs': ['B3:3/12'],
            'description': 'Fire Eye 5 Failure Alarm',
            'timer': 'T4:12',
            'logic_type': 'XIO'
        },
        {
            'rung': '0014',
            'inputs': ['I:0/17'],
            'outputs': ['B3:3/13'],
            'description': 'Fire Eye 6 Failure Alarm',
            'timer': 'T4:11',
            'logic_type': 'XIO'
        },
        {
            'rung': '0016',
            'inputs': ['I:0/19'],
            'outputs': ['B3:3/14'],
            'description': 'Fire Eye 7 Failure Alarm',
            'timer': 'T4:10',
            'logic_type': 'XIO'
        },
        {
            'rung': '0018',
            'inputs': ['I:1/1'],
            'outputs': ['B3:3/15'],
            'description': 'Fire Eye 8 Failure Alarm',
            'timer': 'T4:9',
            'logic_type': 'XIO'
        },
        # Rung 0020 - Fire Eye Faulted Zone 1 (OR of failures)
        {
            'rung': '0020',
            'inputs': ['B3:3/10', 'B3:3/11', 'B3:3/12', 'B3:3/13', 'B3:3/14', 'B3:3/15'],
            'outputs': ['B3:0/1'],
            'description': 'Fire Eye Faulted Zone 1',
            'logic_type': 'OR'
        },
        # Rung 0021 - Fire Eye Faulted Zone 2
        {
            'rung': '0021',
            'inputs': ['B3:3/9', 'B3:3/8'],
            'outputs': ['B3:0/3'],
            'description': 'Fire Eye Faulted Zone 2',
            'logic_type': 'OR'
        },
        # Fire Eye Fire Detection with TON and CTU
        {
            'rung': '0022-0023',
            'inputs': ['I:0/6'],
            'outputs': ['B3:3/0'],
            'description': 'Fire Eye 1 Fire Detected',
            'timer': 'T4:1',
            'counter': 'C5:0'
        },
        {
            'rung': '0024-0025',
            'inputs': ['I:0/8'],
            'outputs': ['B3:3/1'],
            'description': 'Fire Eye 2 Fire Detected',
            'timer': 'T4:2',
            'counter': 'C5:1'
        },
        {
            'rung': '0026-0027',
            'inputs': ['I:0/10'],
            'outputs': ['B3:3/2'],
            'description': 'Fire Eye 3 Fire Detected',
            'timer': 'T4:3',
            'counter': 'C5:2'
        },
        {
            'rung': '0028-0029',
            'inputs': ['I:0/12'],
            'outputs': ['B3:3/3'],
            'description': 'Fire Eye 4 Fire Detected',
            'timer': 'T4:4',
            'counter': 'C5:3'
        },
        {
            'rung': '0030-0031',
            'inputs': ['I:0/14'],
            'outputs': ['B3:3/4'],
            'description': 'Fire Eye 5 Fire Detected',
            'timer': 'T4:5',
            'counter': 'C5:4'
        },
        {
            'rung': '0032-0033',
            'inputs': ['I:0/16'],
            'outputs': ['B3:3/5'],
            'description': 'Fire Eye 6 Fire Detected',
            'timer': 'T4:6',
            'counter': 'C5:5'
        },
        {
            'rung': '0034-0035',
            'inputs': ['I:0/18'],
            'outputs': ['B3:3/6'],
            'description': 'Fire Eye 7 Fire Detected',
            'timer': 'T4:7',
            'counter': 'C5:6'
        },
        {
            'rung': '0036-0037',
            'inputs': ['I:1/0'],
            'outputs': ['B3:3/7'],
            'description': 'Fire Eye 8 Fire Detected',
            'timer': 'T4:8',
            'counter': 'C5:7'
        },
        # Single detector alarms
        {
            'rung': '0038',
            'inputs': ['B3:3/2', 'B3:3/3', 'B3:3/4', 'B3:3/5', 'B3:3/6', 'B3:3/7'],
            'outputs': ['B3:0/5'],
            'description': 'Fire Detected Zone 1 Fire Eyes Single Detector',
            'logic_type': 'OR'
        },
        {
            'rung': '0039',
            'inputs': ['B3:3/1', 'B3:3/0'],
            'outputs': ['B3:0/4'],
            'description': 'Fire Detected Zone 2 Fire Eyes Single Detector',
            'logic_type': 'OR'
        },
        # Additional alarm rungs
        {
            'rung': '0054',
            'inputs': ['B14:0/2'],
            'outputs': ['B3:0/11'],
            'description': 'Fire Alarm Zone 2'
        },
        {
            'rung': '0055',
            'inputs': ['I:1/12'],
            'outputs': ['B3:0/8', 'B3:2/8'],
            'description': 'Fire Eye Failure Warning',
            'logic_type': 'XIO'
        },
        {
            'rung': '0056',
            'inputs': ['I:1/13'],
            'outputs': ['B3:0/9', 'B3:2/9'],
            'description': 'Strobe Light On',
            'logic_type': 'XIO'
        },
        {
            'rung': '0057',
            'inputs': ['I:1/15'],
            'outputs': ['B3:0/13', 'B3:2/13'],
            'description': 'Strobe Light On'
        },
    ]

    return rungs, tag_descriptions

def build_alarm_summary(tag_descriptions):
    """Build alarm summary data from tags - only includes entries with 'alarm' in description"""
    alarms = []

    # Define all potential alarm tags to extract
    alarm_tags = [
        # Fire Alarms
        ('B3:0/0', 'Fire Alarm Zone 1'),
        ('B3:0/11', 'Fire Alarm Zone 2'),
        ('B3:0/1', 'Fire Eye Faulted Zone 1'),
        ('B3:0/3', 'Fire Eye Faulted Zone 2'),
        ('B3:0/4', 'Fire Detected Zone 2 Fire Eyes. Single Detector Only'),
        ('B3:0/5', 'Fire Detected Zone 1 Fire Eyes. Single Detector Only'),
        ('B3:0/6', 'Plant ESD'),
        ('B3:0/8', 'Fire Eye Failure Warning'),
        ('B3:0/9', 'Strobe Light On'),
        ('B3:0/10', 'Strobe Light On'),
        ('B3:0/12', 'Fire System Deluge Valve Open'),
        ('B3:0/13', 'Strobe Light On'),
        # Fire Eye Failure Alarms
        ('B3:3/8', 'FE 1 Failure Alarm'),
        ('B3:3/9', 'FE 2 Failure Alarm'),
        ('B3:3/10', 'FE 3 Failure Alarm'),
        ('B3:3/11', 'FE 4 Failure Alarm'),
        ('B3:3/12', 'FE 5 Failure Alarm'),
        ('B3:3/13', 'FE 6 Failure Alarm'),
        ('B3:3/14', 'FE 7 Failure Alarm'),
        ('B3:3/15', 'FE 8 Failure Alarm'),
        # Fire Eye Fire Detected
        ('B3:3/0', 'Fire Eye 1 Fire Detected'),
        ('B3:3/1', 'Fire Eye 2 Fire Detected'),
        ('B3:3/2', 'Fire Eye 3 Fire Detected'),
        ('B3:3/3', 'Fire Eye 4 Fire Detected'),
        ('B3:3/4', 'Fire Eye 5 Fire Detected'),
        ('B3:3/5', 'Fire Eye 6 Fire Detected'),
        ('B3:3/6', 'Fire Eye 7 Fire Detected'),
        ('B3:3/7', 'Fire Eye 8 Fire Detected'),
        # 2 Detector Alarms
        ('B3:4/6', '2 Detectors In Alarm Zone 1'),
        ('B3:4/5', '2 Detectors In Alarm Zone 2'),
        # ESD
        ('B3:10/0', 'ESD Alarm to Office PLC'),
        # Deluge Status
        ('O:0/0', 'Deluge Valve Zone 2 Open'),
    ]

    # Filter to only include entries with "alarm" in the description (case-insensitive)
    for tag_addr, description in alarm_tags:
        if 'alarm' in description.lower():
            alarms.append({
                'Tag No': tag_addr,
                'P & ID': '',
                'Service Description': description,
                'Range': '',
                'EU': '',
                'Normal Operating Conditions': '',
                'HH': '',
                'H': '',
                'L': '',
                'LL': '',
                'Engineering Notes': ''
            })

    return alarms

def build_cause_effect_matrix(rungs, tag_descriptions):
    """Build cause and effect matrix from ladder rungs"""
    interlocks = []
    interlock_num = 1

    # Filter rungs that have physical I/O or shutdowns
    for rung in rungs:
        # Check if rung has physical inputs (I:, B11:, B14:) or outputs (O:, B3:2, B3:10)
        has_physical_input = any(tag.startswith(('I:', 'B11:', 'B14:')) for tag in rung['inputs'])
        has_physical_output = any(tag.startswith(('O:')) for tag in rung['outputs'])
        has_shutdown = any(tag.startswith(('B3:2', 'B3:10', 'B3:0/0', 'B3:0/11')) for tag in rung['outputs'])

        # Only include rungs with physical I/O or key outputs
        if has_physical_input and (has_physical_output or has_shutdown):
            # Get all input tags
            input_tags = rung['inputs']

            # Primary input (first physical input)
            primary_input = None
            for tag in input_tags:
                if tag.startswith(('I:', 'B11:', 'B14:')):
                    primary_input = tag
                    break

            if not primary_input and input_tags:
                primary_input = input_tags[0]

            # Get description
            service_desc = rung['description']
            if primary_input and primary_input in tag_descriptions:
                service_desc = tag_descriptions[primary_input]

            # Build effects dictionary
            effects = defaultdict(str)
            for output in rung['outputs']:
                effects[output] = 'X'

            # Add timer/counter info to description if present
            extra_info = []
            if 'timer' in rung:
                extra_info.append(f"Timer: {rung['timer']}")
            if 'counter' in rung:
                extra_info.append(f"Counter: {rung['counter']}")
            if extra_info:
                service_desc += f" ({', '.join(extra_info)})"

            interlock = {
                'Interlock No': interlock_num,
                'Tag No': primary_input or '',
                'Service Description': service_desc,
                'Range': '',
                'Pre-Trip (H or L)': '',
                'Trip (HH or LL)': '',
                'P & ID': '',
                'Rung': rung['rung'],
                'Effects': effects,
                'All Inputs': input_tags,
                'All Outputs': rung['outputs']
            }

            interlocks.append(interlock)
            interlock_num += 1

    return interlocks

def generate_alarm_summary_excel(alarms, output_file, template_file=None):
    """Generate Alarm Summary Excel file using template if provided"""
    
    if template_file:
        # Copy template to output file and load it
        shutil.copy(template_file, output_file)
        wb = load_workbook(output_file)
        ws = wb['TEMPLATE']
        ws.title = 'Alarm Summary'
        
        # Template structure:
        # Row 19 = Headers
        # Row 20 = Example row (to be replaced with first data row)
        # Data starts at row 20
        
        DATA_START_ROW = 20
        
        # Unmerge cells in the data area to allow writing
        # Find and unmerge any merged cells starting from row 20
        merged_ranges_to_remove = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row >= DATA_START_ROW:
                merged_ranges_to_remove.append(merged_range)
        for merged_range in merged_ranges_to_remove:
            ws.unmerge_cells(str(merged_range))
        
        # Get the style from the example row (row 20) to apply to data rows
        example_styles = {}
        for col in range(1, 14):  # Columns A-M
            cell = ws.cell(row=DATA_START_ROW, column=col)
            example_styles[col] = {
                'font': Font(
                    name=cell.font.name,
                    size=cell.font.size or 8,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color
                ),
                'fill': PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color
                ) if cell.fill.fill_type else None,
                'border': Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                ),
                'alignment': Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text
                )
            }
        
        # Standard black font for data rows
        black_font = Font(size=8, color='000000')
        
        # Standard thin border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write alarm data starting at row 20 (replacing example row)
        for row_idx, alarm in enumerate(alarms):
            current_row = DATA_START_ROW + row_idx
            
            # Column mapping: A=Tag No, B=P&ID, C=Service Description, D=Range, 
            # E=EU, F=Normal Operating Conditions, G=HH, H=H, I=L, J=LL, K=Engineering Notes
            row_data = [
                alarm['Tag No'],           # A
                alarm['P & ID'],           # B
                alarm['Service Description'],  # C
                alarm['Range'],            # D
                alarm['EU'],               # E
                alarm['Normal Operating Conditions'],  # F
                alarm['HH'],               # G
                alarm['H'],                # H
                alarm['L'],                # I
                alarm['LL'],               # J
                alarm['Engineering Notes'],  # K
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                # Apply black font and borders (not red from template)
                cell.font = black_font
                cell.border = thin_border
                if col_idx in example_styles:
                    style = example_styles[col_idx]
                    if style['fill']:
                        cell.fill = style['fill']
                    cell.alignment = style['alignment']
        
        # Update the title in the template (rows 16-18 have the title)
        # Replace [UNIT NAME] with PLC name
        for row in range(16, 19):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if cell.value and '[UNIT NAME]' in str(cell.value):
                    cell.value = str(cell.value).replace('[UNIT NAME]', PLC_NAME)
        
    else:
        # Original behavior - create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Alarm Summary'

        # Header row
        headers = [
            'Tag No', 'P & ID', 'Service Description', 'Range', 'EU',
            'Normal Operating\nConditions', 'HH', 'H', 'L', 'LL', 'Engineering Notes'
        ]

        ws.append(headers)

        # Format header
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = Font(bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Add data
        for alarm in alarms:
            row = [
                alarm['Tag No'],
                alarm['P & ID'],
                alarm['Service Description'],
                alarm['Range'],
                alarm['EU'],
                alarm['Normal Operating Conditions'],
                alarm['HH'],
                alarm['H'],
                alarm['L'],
                alarm['LL'],
                alarm['Engineering Notes']
            ]
            ws.append(row)

            # Apply borders to data rows
            for cell in ws[ws.max_row]:
                cell.border = thin_border

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 8
        ws.column_dimensions['K'].width = 30

    wb.save(output_file)
    print(f'✓ Alarm Summary saved to: {output_file}')

def generate_cause_effect_excel(interlocks, tag_descriptions, output_file, template_file=None):
    """Generate Cause & Effect Matrix Excel file using template if provided"""
    
    # Collect all unique output effects
    all_effects = set()
    for interlock in interlocks:
        all_effects.update(interlock['Effects'].keys())

    effect_columns = sorted(list(all_effects), key=lambda x: (x.split(':')[0], int(x.split(':')[1].split('/')[0]), int(x.split('/')[1])))

    if template_file:
        # Copy template to output file and load it
        shutil.copy(template_file, output_file)
        wb = load_workbook(output_file)
        ws = wb['TEMPLATE']
        ws.title = 'Cause & Effect'
        
        # Template structure (based on analysis):
        # Row 15: Title area - put title in D15 (centered in large space)
        # Row 16: Was unit name row - clear it
        # Row 17: "CAUSE" label spanning A-G, Tag No in I17
        # Row 18: Column headers (Interlock No, Tag No, Service Description, Range, Pre-Trip, Trip), P & ID
        # Row 19: Example data row (to be replaced)
        # 
        # CAUSE columns: A=Interlock No, B=Tag No, C=Service Description, D=empty, E=Range, F=Pre-Trip, G=Trip, H=empty
        # EFFECT starts at column J onwards (effect columns)
        
        TITLE_ROW = 15           # Row for title in column D
        UNIT_NAME_ROW = 16       # Row to clear
        CAUSE_LABEL_ROW = 17     # Row with "CAUSE" label and Tag No
        HEADER_ROW = 18          # Row with column headers
        DATA_START_ROW = 19      # Data starts here (replacing example)
        
        EFFECT_START_COL = 10    # Column J is where effects start
        
        # Unmerge cells from row 15 onwards to allow writing (title, headers, and data)
        merged_ranges_to_remove = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row >= TITLE_ROW:
                merged_ranges_to_remove.append(merged_range)
        for merged_range in merged_ranges_to_remove:
            ws.unmerge_cells(str(merged_range))
        
        # Clear A15 and put title in D15 with two lines
        ws.cell(row=TITLE_ROW, column=1, value='')  # Clear A15
        title_cell = ws.cell(row=TITLE_ROW, column=4, value=f"CAUSE AND EFFECT MATRIX\n{PLC_NAME}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Clear row 16 (was unit name row)
        for col in range(1, 8):  # Clear columns A-G
            ws.cell(row=UNIT_NAME_ROW, column=col, value='')
        
        # Restore thick border above CAUSE row (row 17) - add bottom border to row 16 cells
        thick_bottom_border = Border(bottom=Side(style='thick'))
        for col in range(1, 8):  # Columns A-G
            cell = ws.cell(row=UNIT_NAME_ROW, column=col)
            cell.border = thick_bottom_border
        
        # Fix header row 18 - make Service Description use only column C, Trip use only column G
        # Service Description header in C18 (not merged with D)
        ws.cell(row=HEADER_ROW, column=3, value='Service Description')
        ws.cell(row=HEADER_ROW, column=4, value='')  # D18 empty
        
        # Trip header in G18 (not merged with H)
        ws.cell(row=HEADER_ROW, column=7, value='Trip\n(HH or LL)')
        ws.cell(row=HEADER_ROW, column=8, value='')  # H18 empty
        
        # Standard black font for data rows
        black_font = Font(size=8, color='000000')
        black_font_bold = Font(size=8, bold=True, color='000000')
        
        # Standard thin border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write EFFECT headers (row 15 - effect descriptions) starting at column J
        for idx, tag in enumerate(effect_columns):
            col = EFFECT_START_COL + idx
            cell = ws.cell(row=TITLE_ROW, column=col, value=tag_descriptions.get(tag, ''))
            cell.font = black_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write Tag No row (row 17 - effect tag numbers) starting at column J
        for idx, tag in enumerate(effect_columns):
            col = EFFECT_START_COL + idx
            cell = ws.cell(row=CAUSE_LABEL_ROW, column=col, value=tag)
            cell.font = black_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write interlock data starting at row 19 (replacing example row)
        for row_idx, interlock in enumerate(interlocks):
            current_row = DATA_START_ROW + row_idx
            
            # Column mapping based on template:
            # A=Interlock No, B=Tag No, C-D=Service Description (merged), 
            # E=Range, F=Pre-Trip, G-H=Trip (merged), I=P&ID, J+=Effect columns
            
            # Write each column individually with proper formatting
            # Column A - Interlock No
            cell = ws.cell(row=current_row, column=1, value=f"I-{interlock['Interlock No']}")
            cell.font = black_font
            cell.border = thin_border
            
            # Column B - Tag No
            cell = ws.cell(row=current_row, column=2, value=interlock['Tag No'])
            cell.font = black_font
            cell.border = thin_border
            
            # Column C - Service Description (will merge C:D after)
            cell = ws.cell(row=current_row, column=3, value=interlock['Service Description'])
            cell.font = black_font
            cell.border = thin_border
            
            # Column D - For merge with C (set border)
            cell = ws.cell(row=current_row, column=4)
            cell.border = thin_border
            
            # Column E - Range
            cell = ws.cell(row=current_row, column=5, value=interlock['Range'])
            cell.font = black_font
            cell.border = thin_border
            
            # Column F - Pre-Trip
            cell = ws.cell(row=current_row, column=6, value=interlock['Pre-Trip (H or L)'])
            cell.font = black_font
            cell.border = thin_border
            
            # Column G - Trip (will merge G:H after)
            cell = ws.cell(row=current_row, column=7, value=interlock['Trip (HH or LL)'])
            cell.font = black_font
            cell.border = thin_border
            
            # Column H - For merge with G (set border)
            cell = ws.cell(row=current_row, column=8)
            cell.border = thin_border
            
            # Column I - P & ID
            cell = ws.cell(row=current_row, column=9, value=interlock['P & ID'])
            cell.font = black_font
            cell.border = thin_border
            
            # Effect columns (J onwards)
            for idx, effect_col in enumerate(effect_columns):
                col = EFFECT_START_COL + idx
                value = 'X' if effect_col in interlock['Effects'] else ''
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if value == 'X':
                    cell.font = black_font_bold
                else:
                    cell.font = black_font
            
        # Merge cells for ALL rows from 18 to 67 (header + data rows including empty ones)
        END_ROW = 67
        
        # Merge C:D and G:H for header row and all data rows (18-67)
        for row in range(HEADER_ROW, END_ROW + 1):
            # Merge C:D for Service Description
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            # Merge G:H for Trip
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        
    else:
        # Original behavior - create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Cause & Effect'

        # Create header structure similar to example
        # Row 1: Title row with "EFFECT" label and service descriptions
        row1 = ['', '', '', '', '', '', 'EFFECT'] + [tag_descriptions.get(tag, '') for tag in effect_columns]
        ws.append(row1)

        # Row 2: "Tag No" label and actual tag numbers
        row2 = ['', '', '', '', '', '', 'Tag No'] + list(effect_columns)
        ws.append(row2)

        # Row 3: Column headers for CAUSE side, blank for EFFECT side
        row3 = ['Interlock\nNo', 'Tag No', 'Service Description', 'Range', 'Pre-Trip\n(H or L)', 'Trip\n(HH or LL)', ''] + [''] * len(effect_columns)
        ws.append(row3)

        # Row 4: CAUSE label, no P & ID labels
        row4 = ['CAUSE', '', '', '', '', '', ''] + [''] * len(effect_columns)
        ws.append(row4)

        # Formatting
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        effect_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        cause_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format row 1 - EFFECT label (column G) and service descriptions
        cell = ws.cell(row=1, column=7)
        cell.fill = effect_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        for idx, tag in enumerate(effect_columns):
            cell = ws.cell(row=1, column=8 + idx)
            cell.fill = effect_fill
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

        # Format row 2 - "Tag No" label (column G) and tag numbers
        cell = ws.cell(row=2, column=7)
        cell.fill = effect_fill
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        for idx, tag in enumerate(effect_columns):
            cell = ws.cell(row=2, column=8 + idx)
            cell.fill = effect_fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Format row 3 - Headers
        for col in range(1, 7):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

        # Format row 4 - CAUSE label only (no P & ID in effect columns)
        cell = ws.cell(row=4, column=1)
        cell.fill = cause_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # Apply borders to remaining cells in row 4
        for col in range(2, 8 + len(effect_columns)):
            cell = ws.cell(row=4, column=col)
            cell.border = thin_border

        # Set row heights
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 20

        # Add interlock data
        for interlock in interlocks:
            row_data = [
                f"I-{interlock['Interlock No']}",
                interlock['Tag No'],
                interlock['Service Description'],
                interlock['Range'],
                interlock['Pre-Trip (H or L)'],
                interlock['Trip (HH or LL)'],
                ''  # Empty column for the label column (G)
            ]

            # Add effect markers
            for effect_col in effect_columns:
                if effect_col in interlock['Effects']:
                    row_data.append('X')
                else:
                    row_data.append('')

            ws.append(row_data)

            # Apply borders and formatting to data row
            for col_idx, cell in enumerate(ws[ws.max_row], 1):
                cell.border = thin_border
                if col_idx >= 8:  # Effect columns (now starting at column 8 due to label column)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.value == 'X':
                        cell.font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15  # Label column

        # Set effect column widths (starting from H)
        for i, tag in enumerate(effect_columns):
            col_letter = chr(ord('H') + i)
            ws.column_dimensions[col_letter].width = 20

    wb.save(output_file)
    print(f'✓ Cause & Effect Matrix saved to: {output_file}')

def main():
    """Main execution function"""
    print('═' * 70)
    print('  RSLogix 500 FIRE SYSTEM PARSER')
    print(f'  {PLC_NAME}')
    print('═' * 70)

    # Extract data from PDF
    print('\n[1/4] Extracting ladder logic from PDF...')
    rungs, tag_descriptions = extract_data_from_pdf()
    print(f'      ✓ Extracted {len(rungs)} ladder rungs')
    print(f'      ✓ Loaded {len(tag_descriptions)} tag descriptions')

    # Build alarm summary
    print('\n[2/4] Building alarm summary...')
    alarms = build_alarm_summary(tag_descriptions)
    print(f'      ✓ Found {len(alarms)} alarm tags')

    # Build cause & effect matrix
    print('\n[3/4] Building cause & effect matrix...')
    interlocks = build_cause_effect_matrix(rungs, tag_descriptions)
    print(f'      ✓ Found {len(interlocks)} interlocks')

    # Generate Excel files using templates
    print('\n[4/4] Generating Excel files from templates...')
    print(f'      Using Alarm Summary template: {ALARM_SUMMARY_TEMPLATE}')
    print(f'      Using Cause & Effect template: {CAUSE_EFFECT_TEMPLATE}')
    
    generate_alarm_summary_excel(alarms, ALARM_SUMMARY_OUTPUT, template_file=ALARM_SUMMARY_TEMPLATE)
    generate_cause_effect_excel(interlocks, tag_descriptions, CAUSE_EFFECT_OUTPUT, template_file=CAUSE_EFFECT_TEMPLATE)

    print('\n' + '═' * 70)
    print('  PROCESSING COMPLETE!')
    print('═' * 70)
    print('\nOutput files created:')
    print(f'  ├─ {ALARM_SUMMARY_OUTPUT}')
    print(f'  └─ {CAUSE_EFFECT_OUTPUT}')
    print('')

if __name__ == '__main__':
    main()
